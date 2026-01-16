import pandas as pd
import openpyxl

# 读取模板文件
template_file = 'excel/模板.xlsx'
wb = openpyxl.load_workbook(template_file)

print("=== 模板文件信息 ===")
print(f"工作表列表: {wb.sheetnames}")
print()

# 查看每个工作表的内容
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- 工作表: {sheet_name} ---")
    print(f"数据范围: {ws.dimensions}")

    # 读取前几行数据
    df = pd.read_excel(template_file, sheet_name=sheet_name)
    print(f"数据形状: {df.shape}")
    print(f"列名: {df.columns.tolist()}")
    print("\n前5行数据:")
    print(df.head())
    print()
